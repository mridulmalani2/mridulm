"""Excel export engine -- 8-sheet professional PE workbook."""

from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from backend.models.state import ModelState

# ── Design System ────────────────────────────────────────────────────────

NAVY_HEX = "1a2744"
LIGHT_GREY_HEX = "f5f7fa"
MID_GREY_HEX = "e8ebf0"
BORDER_GREY_HEX = "d0d5dd"
GREEN_HEX = "006100"
GREEN_BG_HEX = "e6f4ea"
RED_HEX = "9C0006"
RED_BG_HEX = "fce8e6"
AMBER_HEX = "9C6500"
AMBER_BG_HEX = "fef7e0"
ACCENT_RED_HEX = "CC0000"

# Fills
NAVY_FILL = PatternFill(start_color=NAVY_HEX, end_color=NAVY_HEX, fill_type="solid")
LIGHT_FILL = PatternFill(start_color=LIGHT_GREY_HEX, end_color=LIGHT_GREY_HEX, fill_type="solid")
MID_FILL = PatternFill(start_color=MID_GREY_HEX, end_color=MID_GREY_HEX, fill_type="solid")
GREEN_BG_FILL = PatternFill(start_color=GREEN_BG_HEX, end_color=GREEN_BG_HEX, fill_type="solid")
RED_BG_FILL = PatternFill(start_color=RED_BG_HEX, end_color=RED_BG_HEX, fill_type="solid")
AMBER_BG_FILL = PatternFill(start_color=AMBER_BG_HEX, end_color=AMBER_BG_HEX, fill_type="solid")
ACCENT_FILL = PatternFill(start_color=ACCENT_RED_HEX, end_color=ACCENT_RED_HEX, fill_type="solid")

# Fonts
F_TITLE = Font(name="Calibri", size=18, bold=True, color=NAVY_HEX)
F_SUBTITLE = Font(name="Calibri", size=11, color="666666")
F_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
F_SECTION = Font(name="Calibri", size=11, bold=True, color=NAVY_HEX)
F_BODY = Font(name="Calibri", size=10, color="111111")
F_BODY_BOLD = Font(name="Calibri", size=10, bold=True, color="111111")
F_INPUT = Font(name="Calibri", size=10, color="0000FF")
F_GREEN = Font(name="Calibri", size=10, bold=True, color=GREEN_HEX)
F_RED = Font(name="Calibri", size=10, bold=True, color=RED_HEX)
F_AMBER = Font(name="Calibri", size=10, bold=True, color=AMBER_HEX)
F_COVER_METRIC = Font(name="Calibri", size=22, bold=True, color=NAVY_HEX)
F_COVER_LABEL = Font(name="Calibri", size=9, color="888888")
F_WHITE = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
F_WHITE_LG = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
F_SMALL = Font(name="Calibri", size=8, color="999999")
F_SM_BOLD = Font(name="Calibri", size=9, color="111111")

# Borders
THIN_BOTTOM = Border(bottom=Side(style="thin", color=BORDER_GREY_HEX))
MED_BOTTOM = Border(bottom=Side(style="medium", color=NAVY_HEX))
THICK_BOTTOM = Border(bottom=Side(style="thick", color=NAVY_HEX))
DOUBLE_TOP_BOTTOM = Border(
    top=Side(style="medium", color=NAVY_HEX),
    bottom=Side(style="double", color=NAVY_HEX),
)

# Number formats
FMT_CCY = '#,##0.0;(#,##0.0);"-"'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_MULT = '0.0"x"'
FMT_NUM = '#,##0.0;(#,##0.0);"-"'
FMT_INT = "#,##0"

CURRENCY_SYMBOLS = {"GBP": "\u00a3", "EUR": "\u20ac", "USD": "$", "CHF": "CHF "}


def _irr_font(irr: Optional[float]) -> Font:
    if irr is None:
        return F_BODY
    if irr > 0.25:
        return F_GREEN
    if irr >= 0.15:
        return F_AMBER
    return F_RED


def _irr_bg(irr: Optional[float]):
    if irr is None:
        return None
    if irr > 0.25:
        return GREEN_BG_FILL
    if irr >= 0.15:
        return AMBER_BG_FILL
    return RED_BG_FILL


def _verdict_font(v: str) -> Font:
    if v == "aggressive":
        return F_RED
    if v == "conservative":
        return F_AMBER
    return F_GREEN


def _verdict_fill(v: str):
    if v == "aggressive":
        return PatternFill(start_color=RED_HEX, end_color=RED_HEX, fill_type="solid")
    if v == "conservative":
        return PatternFill(start_color=AMBER_HEX, end_color=AMBER_HEX, fill_type="solid")
    return PatternFill(start_color=GREEN_HEX, end_color=GREEN_HEX, fill_type="solid")


# ── Helpers ──────────────────────────────────────────────────────────────

def _style_header_row(ws, row: int, from_col: int, to_col: int):
    for col in range(from_col, to_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = NAVY_FILL
        cell.font = F_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BOTTOM


def _write_section_header(ws, row: int, title: str, max_col: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = F_SECTION
    cell.border = MED_BOTTOM
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22
    return row + 1


def _write_data_row(ws, row: int, label: str, values: list, fmt: str = FMT_CCY,
                    alt: bool = False, bold: bool = False, top_border: bool = False) -> int:
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = F_BODY_BOLD if bold else F_BODY
    lc.border = THIN_BOTTOM
    if alt:
        lc.fill = LIGHT_FILL

    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=i + 2, value=v)
        cell.font = F_BODY_BOLD if bold else F_BODY
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right")
        if top_border:
            cell.border = Border(
                top=Side(style="medium", color=NAVY_HEX),
                bottom=Side(style="thin", color=BORDER_GREY_HEX),
            )
        else:
            cell.border = THIN_BOTTOM
        if alt:
            cell.fill = LIGHT_FILL
    return row + 1


def _write_kv_row(ws, row: int, label: str, value, fmt: str = None,
                  font=None, alt: bool = False, is_input: bool = False) -> int:
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = F_BODY
    lc.border = THIN_BOTTOM
    if alt:
        lc.fill = LIGHT_FILL

    vc = ws.cell(row=row, column=2, value=value)
    vc.font = F_INPUT if is_input else (font or F_BODY)
    if fmt:
        vc.number_format = fmt
    elif isinstance(value, float) and abs(value) < 1 and value != 0:
        vc.number_format = FMT_PCT
    elif isinstance(value, (int, float)):
        vc.number_format = FMT_CCY
    vc.alignment = Alignment(horizontal="right")
    vc.border = THIN_BOTTOM
    if alt:
        vc.fill = LIGHT_FILL
    return row + 1


def _set_print(ws, landscape: bool = True):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = 9
    ws.oddFooter.left.text = "Deal Intelligence Engine"
    ws.oddFooter.left.size = 8
    ws.oddFooter.center.text = "Confidential"
    ws.oddFooter.center.size = 8
    ws.oddFooter.right.text = "Page &P of &N"
    ws.oddFooter.right.size = 8


# ── Main Entry Point ─────────────────────────────────────────────────────

def build_excel(state: ModelState) -> bytes:
    """Build an 8-sheet Excel workbook and return as bytes."""
    wb = Workbook()
    ccy = CURRENCY_SYMBOLS.get(state.currency, "\u00a3")
    hp = state.exit.holding_period

    _build_cover_sheet(wb, state, ccy)
    _build_sources_uses_sheet(wb, state, ccy)
    _build_assumptions_sheet(wb, state, ccy)
    _build_pl_sheet(wb, state, ccy, hp)
    _build_cash_flow_debt_sheet(wb, state, ccy, hp)
    _build_returns_sheet(wb, state, ccy, hp)
    _build_scenarios_sheet(wb, state, ccy)
    _build_risk_sheet(wb, state, ccy, hp)

    # Remove default empty sheet if it exists and is unused
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Sheet 1: Cover ───────────────────────────────────────────────────────

def _build_cover_sheet(wb: Workbook, state: ModelState, ccy: str):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_properties.tabColor = ACCENT_RED_HEX
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 4

    # Accent bar
    for c in range(1, 8):
        ws.cell(row=1, column=c).fill = ACCENT_FILL
        ws.cell(row=2, column=c).fill = ACCENT_FILL
    ws.row_dimensions[1].height = 4
    ws.row_dimensions[2].height = 4

    row = 4
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value="DEAL INTELLIGENCE ENGINE").font = Font(
        name="Calibri", size=10, bold=True, color="999999"
    )
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value=state.deal_name).font = F_TITLE
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(
        row=row, column=2,
        value=f"{state.sector}  |  {state.currency}  |  {state.exit.holding_period}-Year Hold",
    ).font = F_SUBTITLE
    row += 2

    for c in range(2, 7):
        ws.cell(row=row, column=c).border = Border(top=Side(style="medium", color=NAVY_HEX))
    row += 1

    ret = state.returns
    metrics = [
        ("EQUITY IRR", ret.irr, FMT_PCT),
        ("MOIC", ret.moic, FMT_MULT),
        ("ENTRY EV", state.entry.enterprise_value, f'{ccy}#,##0.0"m"'),
        ("ENTRY EQUITY", ret.entry_equity, f'{ccy}#,##0.0"m"'),
        ("EXIT EV", ret.exit_ev, f'{ccy}#,##0.0"m"'),
        ("LEVERAGE", state.entry.leverage_ratio, FMT_MULT),
    ]
    cols = [2, 4, 6]
    for i, (label, value, fmt) in enumerate(metrics):
        c = cols[i % 3]
        r = row + (i // 3) * 3
        vc = ws.cell(row=r, column=c, value=value)
        vc.font = F_COVER_METRIC
        if isinstance(value, (int, float)):
            vc.number_format = fmt
        if label == "EQUITY IRR" and isinstance(value, float):
            vc.font = Font(name="Calibri", size=22, bold=True, color=_irr_font(value).color)
        ws.cell(row=r + 1, column=c, value=label).font = F_COVER_LABEL

    row += 7
    for c in range(2, 7):
        ws.cell(row=row, column=c).border = Border(top=Side(style="thin", color=BORDER_GREY_HEX))
    row += 1

    # Scenario overview
    if state.scenarios:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row=row, column=2, value="SCENARIO OVERVIEW").font = F_SECTION
        ws.cell(row=row, column=2).border = MED_BOTTOM
        row += 1
        headers = [""] + [s.name.upper() for s in state.scenarios]
        for i, h in enumerate(headers):
            ws.cell(row=row, column=i + 2, value=h)
        _style_header_row(ws, row, 2, len(headers) + 1)
        row += 1

        ws.cell(row=row, column=2, value="IRR").font = F_BODY
        for i, s in enumerate(state.scenarios):
            c = ws.cell(row=row, column=i + 3, value=s.irr)
            c.number_format = FMT_PCT
            c.font = _irr_font(s.irr)
            bg = _irr_bg(s.irr)
            if bg:
                c.fill = bg
        row += 1

        ws.cell(row=row, column=2, value="MOIC").font = F_BODY
        for i, s in enumerate(state.scenarios):
            c = ws.cell(row=row, column=i + 3, value=s.moic)
            c.number_format = FMT_MULT
        row += 2

    # Reality check
    rc = state.exit_reality_check
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2, value="EXIT REALITY CHECK").font = F_SECTION
    vc = ws.cell(row=row, column=4, value=rc.verdict.upper())
    vc.font = _verdict_font(rc.verdict)
    vc.fill = _verdict_fill(rc.verdict) if rc.verdict == "aggressive" else (
        PatternFill(start_color=AMBER_BG_HEX, end_color=AMBER_BG_HEX, fill_type="solid")
        if rc.verdict == "conservative" else GREEN_BG_FILL
    )
    vc.alignment = Alignment(horizontal="center")
    row += 3

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(
        row=row, column=2,
        value=f"Generated {datetime.utcnow().strftime('%d %B %Y')} | Deal Intelligence Engine | Confidential",
    ).font = F_SMALL

    _set_print(ws, landscape=False)


# ── Sheet 2: Sources & Uses ──────────────────────────────────────────────

def _build_sources_uses_sheet(wb: Workbook, state: ModelState, ccy: str):
    ws = wb.create_sheet("Sources & Uses")
    ws.sheet_properties.tabColor = NAVY_HEX
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14

    su = state.sources_and_uses
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value=f"{state.deal_name} - Sources & Uses ({ccy}m)").font = F_SECTION
    ws.cell(row=row, column=1).border = THICK_BOTTOM
    row += 2

    # Headers
    for col, val in [(1, "USES"), (2, f"{ccy}m"), (3, "% Total"), (5, "SOURCES"), (6, f"{ccy}m"), (7, "% Total")]:
        ws.cell(row=row, column=col, value=val)
    _style_header_row(ws, row, 1, 3)
    _style_header_row(ws, row, 5, 7)
    row += 1

    total = su.total_uses or 1

    def write_use(r, label, amt, alt=False):
        ws.cell(row=r, column=1, value=label).font = F_BODY
        ws.cell(row=r, column=1).border = THIN_BOTTOM
        ws.cell(row=r, column=2, value=amt).font = F_BODY
        ws.cell(row=r, column=2).number_format = FMT_CCY
        ws.cell(row=r, column=2).border = THIN_BOTTOM
        ws.cell(row=r, column=3, value=amt / total).font = F_BODY
        ws.cell(row=r, column=3).number_format = FMT_PCT
        ws.cell(row=r, column=3).border = THIN_BOTTOM
        if alt:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = LIGHT_FILL
        return r + 1

    def write_src(r, label, amt, alt=False):
        ws.cell(row=r, column=5, value=label).font = F_BODY
        ws.cell(row=r, column=5).border = THIN_BOTTOM
        ws.cell(row=r, column=6, value=amt).font = F_BODY
        ws.cell(row=r, column=6).number_format = FMT_CCY
        ws.cell(row=r, column=6).border = THIN_BOTTOM
        ws.cell(row=r, column=7, value=amt / total).font = F_BODY
        ws.cell(row=r, column=7).number_format = FMT_PCT
        ws.cell(row=r, column=7).border = THIN_BOTTOM
        if alt:
            for c in range(5, 8):
                ws.cell(row=r, column=c).fill = LIGHT_FILL
        return r + 1

    ur = row
    ur = write_use(ur, "Enterprise Value", su.enterprise_value)
    ur = write_use(ur, "Transaction Fees", su.transaction_fees, True)
    ur = write_use(ur, "Financing Fees", su.financing_fees)
    # Total uses
    for c, val in [(1, "Total Uses"), (2, su.total_uses), (3, 1.0)]:
        cell = ws.cell(row=ur, column=c, value=val)
        cell.font = F_BODY_BOLD
        cell.border = DOUBLE_TOP_BOTTOM
        if c > 1:
            cell.number_format = FMT_CCY if c == 2 else FMT_PCT

    sr = row
    for i, d in enumerate(su.debt_sources):
        sr = write_src(sr, d.name, d.amount, i % 2 == 1)
    sr = write_src(sr, "Total Debt", su.total_debt)
    if su.rollover_equity > 0:
        sr = write_src(sr, "Rollover Equity", su.rollover_equity, True)
    sr = write_src(sr, "Sponsor Equity", su.sponsor_equity, su.rollover_equity > 0)
    for c, val in [(5, "Total Sources"), (6, su.total_sources), (7, 1.0)]:
        cell = ws.cell(row=sr, column=c, value=val)
        cell.font = F_BODY_BOLD
        cell.border = DOUBLE_TOP_BOTTOM
        if c > 5:
            cell.number_format = FMT_CCY if c == 6 else FMT_PCT

    row = max(ur, sr) + 1

    # ── Audit check: Sources = Uses (hard reconciliation) ──
    recon_delta = su.total_sources - su.total_uses
    balanced = abs(recon_delta) < 0.01
    audit_label = "AUDIT CHECK: Sources = Uses" if balanced else "AUDIT CHECK: Sources \u2260 Uses — RECONCILIATION ERROR"
    audit_font = F_GREEN if balanced else F_RED
    audit_fill = GREEN_BG_FILL if balanced else RED_BG_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    audit_cell = ws.cell(row=row, column=1, value=audit_label)
    audit_cell.font = audit_font
    audit_cell.fill = audit_fill
    audit_cell.border = MED_BOTTOM
    audit_cell.alignment = Alignment(horizontal="left", vertical="center")
    delta_cell = ws.cell(row=row, column=4, value=recon_delta if not balanced else 0.0)
    delta_cell.font = audit_font
    delta_cell.fill = audit_fill
    delta_cell.number_format = FMT_CCY
    delta_cell.border = MED_BOTTOM
    row += 2

    row = _write_section_header(ws, row, "CAPITALISATION METRICS", 2)
    row = _write_kv_row(ws, row, "Implied Leverage (Debt / EBITDA)", su.implied_leverage, fmt=FMT_MULT)
    row = _write_kv_row(ws, row, "Equity as % of Total", su.equity_pct_of_total, fmt=FMT_PCT, alt=True)
    row = _write_kv_row(ws, row, "Debt as % of Total", su.debt_pct_of_total, fmt=FMT_PCT)
    # Breakdown of Uses for auditability
    row += 1
    row = _write_section_header(ws, row, "USES BREAKDOWN (for audit)", 2)
    row = _write_kv_row(ws, row, "Enterprise Value", su.enterprise_value, fmt=FMT_CCY)
    row = _write_kv_row(ws, row, "Transaction & Advisory Fees", su.transaction_fees, fmt=FMT_CCY, alt=True)
    row = _write_kv_row(ws, row, "Financing Fees", su.financing_fees, fmt=FMT_CCY)
    row = _write_kv_row(ws, row, "Total Uses (= Sponsor Equity + Debt)", su.total_uses, fmt=FMT_CCY, alt=True)

    _set_print(ws)


# ── Sheet 3: Assumptions ────────────────────────────────────────────────

def _build_assumptions_sheet(wb: Workbook, state: ModelState, ccy: str):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_properties.tabColor = "1a5276"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    ai = set(state.ai_toggle_fields)
    row = 1

    def add(label, value, field_path="", fmt=None, alt=False):
        nonlocal row
        prefix = "[AI] " if field_path in ai else ""
        row = _write_kv_row(ws, row, f"{prefix}{label}", value, fmt=fmt, alt=alt, is_input=field_path in ai)

    row = _write_section_header(ws, row, "ENTRY ASSUMPTIONS", 2)
    add("Deal Name", state.deal_name)
    add("Sector", state.sector, alt=True)
    add("Currency", state.currency)
    add("LTM Revenue", state.revenue.base_revenue, fmt=FMT_CCY, alt=True)
    add("LTM EBITDA", state.revenue.base_revenue * state.margins.base_ebitda_margin, fmt=FMT_CCY)
    add("LTM EBITDA Margin", state.margins.base_ebitda_margin, "margins.base_ebitda_margin", FMT_PCT, True)
    add("Entry EBITDA Multiple", state.entry.entry_ebitda_multiple, fmt=FMT_MULT)
    add("Enterprise Value", state.entry.enterprise_value, fmt=FMT_CCY, alt=True)
    add("Total Debt Raised", state.entry.total_debt_raised, fmt=FMT_CCY)
    add("Leverage Ratio", state.entry.leverage_ratio, "entry.leverage_ratio", FMT_MULT, True)
    add("Equity Check", state.entry.equity_check, fmt=FMT_CCY)
    row += 1

    row = _write_section_header(ws, row, "DEBT STRUCTURE", 2)
    for t in state.debt_tranches:
        ws.cell(row=row, column=1, value=t.name).font = F_BODY_BOLD
        ws.cell(row=row, column=1).fill = MID_FILL
        ws.cell(row=row, column=2).fill = MID_FILL
        row += 1
        add("  Principal", t.principal, fmt=FMT_CCY)
        add("  Rate Type", t.rate_type, alt=True)
        add("  Interest Rate", t.interest_rate, fmt=FMT_PCT)
        add("  Amortization", t.amortization_type, alt=True)
        if t.pik_rate > 0:
            add("  PIK Rate", t.pik_rate, fmt=FMT_PCT)
    row += 1

    row = _write_section_header(ws, row, "REVENUE ASSUMPTIONS", 2)
    for i, g in enumerate(state.revenue.growth_rates):
        add(f"Year {i + 1} Revenue Growth", g, "revenue.growth_rates", FMT_PCT, i % 2 == 1)
    row += 1

    row = _write_section_header(ws, row, "MARGIN & OPERATING", 2)
    add("Base EBITDA Margin", state.margins.base_ebitda_margin, "margins.base_ebitda_margin", FMT_PCT)
    add("Target EBITDA Margin", state.margins.target_ebitda_margin, "margins.target_ebitda_margin", FMT_PCT, True)
    add("D&A (% Revenue)", state.margins.da_pct_revenue, fmt=FMT_PCT)
    add("Capex (% Revenue)", state.margins.capex_pct_revenue, fmt=FMT_PCT, alt=True)
    add("NWC (% Revenue)", state.margins.nwc_pct_revenue, fmt=FMT_PCT)
    row += 1

    row = _write_section_header(ws, row, "FEES & TAX", 2)
    add("Entry Fee %", state.fees.entry_fee_pct, fmt=FMT_PCT)
    add("Exit Fee %", state.fees.exit_fee_pct, fmt=FMT_PCT, alt=True)
    add("Monitoring Fee (p.a.)", state.fees.monitoring_fee_annual, "fees.monitoring_fee_annual", FMT_CCY)
    add("Financing Fee %", state.fees.financing_fee_pct, fmt=FMT_PCT, alt=True)
    add("Transaction Costs", state.fees.transaction_costs, fmt=FMT_CCY)
    add("Tax Rate", state.tax.tax_rate, fmt=FMT_PCT, alt=True)
    row += 1

    row = _write_section_header(ws, row, "EXIT ASSUMPTIONS", 2)
    add("Holding Period", state.exit.holding_period, "exit.holding_period", FMT_INT)
    add("Exit EBITDA Multiple", state.exit.exit_ebitda_multiple, "exit.exit_ebitda_multiple", FMT_MULT, True)
    add("Exit Method", state.exit.exit_method)
    row += 1

    row = _write_section_header(ws, row, "MANAGEMENT INCENTIVE", 2)
    add("MIP Pool %", state.mip.mip_pool_pct, fmt=FMT_PCT)
    add("Hurdle MOIC", state.mip.hurdle_moic, fmt=FMT_MULT, alt=True)
    add("Vesting Years", state.mip.vesting_years, fmt=FMT_INT)
    add("Sweet Equity %", state.mip.sweet_equity_pct, fmt=FMT_PCT, alt=True)

    _set_print(ws, landscape=False)


# ── Sheet 4: P&L ────────────────────────────────────────────────────────

def _build_pl_sheet(wb: Workbook, state: ModelState, ccy: str, hp: int):
    ws = wb.create_sheet("P&L")
    ws.sheet_properties.tabColor = "2e86c1"
    ws.column_dimensions["A"].width = 30
    for c in range(2, hp + 3):
        ws.column_dimensions[get_column_letter(c)].width = 14

    years = state.projections.years
    if not years:
        ws.cell(row=1, column=1, value="No projections calculated")
        return

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=hp + 2)
    ws.cell(row=row, column=1, value=f"{state.deal_name} - Income Statement ({ccy}m)").font = F_SECTION
    ws.cell(row=row, column=1).border = THICK_BOTTOM
    row += 1

    ws.cell(row=row, column=2, value="Entry (LTM)")
    for i in range(hp):
        ws.cell(row=row, column=i + 3, value=f"Year {i + 1}")
    _style_header_row(ws, row, 1, hp + 2)
    row += 1

    entry_rev = state.revenue.base_revenue
    entry_ebitda = entry_rev * state.margins.base_ebitda_margin

    row = _write_data_row(ws, row, "Revenue", [entry_rev] + [y.revenue for y in years], FMT_CCY, bold=True)
    row = _write_data_row(ws, row, "Revenue Growth", [None] + [y.revenue_growth for y in years], FMT_PCT, alt=True)
    row += 1
    row = _write_data_row(ws, row, "EBITDA", [entry_ebitda] + [y.ebitda for y in years], FMT_CCY, bold=True)
    row = _write_data_row(ws, row, "EBITDA Margin", [state.margins.base_ebitda_margin] + [y.ebitda_margin for y in years], FMT_PCT, alt=True)
    row = _write_data_row(ws, row, "Monitoring Fee Adj.", [0] + [-state.fees.monitoring_fee_annual] * hp, FMT_CCY)
    row = _write_data_row(ws, row, "EBITDA Adjusted", [entry_ebitda] + [y.ebitda_adj for y in years], FMT_CCY, bold=True, alt=True)
    row += 1
    row = _write_data_row(ws, row, "D&A", [None] + [-y.da for y in years], FMT_CCY)
    row = _write_data_row(ws, row, "EBIT", [None] + [y.ebit for y in years], FMT_CCY, bold=True, alt=True)
    row = _write_data_row(ws, row, "Interest Expense", [None] + [-y.interest_expense for y in years], FMT_CCY)
    row = _write_data_row(ws, row, "EBT", [None] + [y.ebt for y in years], FMT_CCY, bold=True)
    row = _write_data_row(ws, row, "Tax", [None] + [-y.tax for y in years], FMT_CCY, alt=True)
    row = _write_data_row(ws, row, "Net Income", [None] + [y.net_income for y in years], FMT_CCY, bold=True, top_border=True)

    _set_print(ws)


# ── Sheet 5: Cash Flow & Debt ───────────────────────────────────────────

def _build_cash_flow_debt_sheet(wb: Workbook, state: ModelState, ccy: str, hp: int):
    ws = wb.create_sheet("Cash Flow & Debt")
    ws.sheet_properties.tabColor = "1e8449"
    ws.column_dimensions["A"].width = 30
    for c in range(2, hp + 2):
        ws.column_dimensions[get_column_letter(c)].width = 14

    years = state.projections.years
    ds = state.debt_schedule
    if not years:
        ws.cell(row=1, column=1, value="No projections calculated")
        return

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=hp + 1)
    ws.cell(row=row, column=1, value=f"{state.deal_name} - Cash Flow & Debt ({ccy}m)").font = F_SECTION
    ws.cell(row=row, column=1).border = THICK_BOTTOM
    row += 1

    for i in range(hp):
        ws.cell(row=row, column=i + 2, value=f"Year {i + 1}")
    _style_header_row(ws, row, 1, hp + 1)
    row += 1

    row = _write_section_header(ws, row, "FREE CASH FLOW BUILD", hp + 1)
    row = _write_data_row(ws, row, "EBITDA Adjusted", [y.ebitda_adj for y in years], FMT_CCY, bold=True)
    row = _write_data_row(ws, row, "Tax Paid", [-y.tax for y in years], FMT_CCY, alt=True)
    row = _write_data_row(ws, row, "Maintenance Capex", [-y.maintenance_capex for y in years], FMT_CCY)
    row = _write_data_row(ws, row, "Growth Capex", [-y.growth_capex for y in years], FMT_CCY, alt=True)
    row = _write_data_row(ws, row, "Change in NWC", [-y.delta_nwc for y in years], FMT_CCY)
    row = _write_data_row(ws, row, "FCF Pre-Debt Service", [y.fcf_pre_debt for y in years], FMT_CCY, bold=True, top_border=True, alt=True)
    row = _write_data_row(ws, row, "Cash Interest", [-v for v in ds.total_cash_interest_by_year], FMT_CCY)
    row = _write_data_row(ws, row, "Debt Repayment", [-v for v in ds.total_repayment_by_year], FMT_CCY, alt=True)
    row = _write_data_row(ws, row, "FCF to Equity", [y.fcf_to_equity for y in years], FMT_CCY, bold=True, top_border=True)
    row += 1

    # Debt schedule
    row = _write_section_header(ws, row, "DEBT SCHEDULE", hp + 1)
    for t_idx, tranche_sched in enumerate(ds.tranche_schedules):
        if not tranche_sched:
            continue
        name = tranche_sched[0].tranche_name if tranche_sched else f"Tranche {t_idx + 1}"
        ws.cell(row=row, column=1, value=name).font = F_BODY_BOLD
        ws.cell(row=row, column=1).fill = MID_FILL
        for c in range(2, hp + 2):
            ws.cell(row=row, column=c).fill = MID_FILL
        row += 1
        row = _write_data_row(ws, row, "  Opening Balance", [y.beginning_balance for y in tranche_sched], FMT_CCY)
        row = _write_data_row(ws, row, "  Cash Interest", [-y.cash_interest for y in tranche_sched], FMT_CCY, alt=True)
        if any(y.pik_accrual > 0 for y in tranche_sched):
            row = _write_data_row(ws, row, "  PIK Accrual", [y.pik_accrual for y in tranche_sched], FMT_CCY)
        row = _write_data_row(ws, row, "  Repayment", [-y.total_repayment for y in tranche_sched], FMT_CCY, alt=True)
        row = _write_data_row(ws, row, "  Closing Balance", [y.ending_balance for y in tranche_sched], FMT_CCY, bold=True, top_border=True)
        row += 1

    row = _write_data_row(ws, row, "Total Debt Outstanding", ds.total_debt_by_year, FMT_CCY, bold=True)
    row += 1

    row = _write_section_header(ws, row, "CREDIT METRICS", hp + 1)
    row = _write_data_row(ws, row, "Net Debt / EBITDA", ds.leverage_ratio_by_year, FMT_MULT, bold=True)
    row = _write_data_row(ws, row, "Interest Coverage", [min(v, 99) for v in ds.interest_coverage_by_year], FMT_NUM, alt=True)
    row = _write_data_row(ws, row, "DSCR", [min(v, 99) for v in ds.dscr_by_year], FMT_NUM)

    ca = state.credit_analysis
    if ca.metrics_by_year:
        row = _write_data_row(ws, row, "FCCR", [min(m.fccr, 99) for m in ca.metrics_by_year], FMT_NUM, alt=True)
        row = _write_data_row(ws, row, "Cumulative Paydown", [m.cumulative_debt_paydown for m in ca.metrics_by_year], FMT_CCY)
        row = _write_data_row(ws, row, "Paydown (% Entry)", [m.debt_paydown_pct for m in ca.metrics_by_year], FMT_PCT, alt=True)

    _set_print(ws)


# ── Sheet 6: Returns ────────────────────────────────────────────────────

def _build_returns_sheet(wb: Workbook, state: ModelState, ccy: str, hp: int):
    ws = wb.create_sheet("Returns")
    ws.sheet_properties.tabColor = "7d3c98"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    ret = state.returns
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"{state.deal_name} - Returns Analysis ({ccy}m)").font = F_SECTION
    ws.cell(row=row, column=1).border = THICK_BOTTOM
    row += 2

    row = _write_section_header(ws, row, "RETURNS SUMMARY", 3)

    def add_ret(label, value, fmt, alt=False):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = F_BODY
        ws.cell(row=row, column=1).border = THIN_BOTTOM
        vc = ws.cell(row=row, column=2, value=value)
        vc.number_format = fmt
        vc.alignment = Alignment(horizontal="right")
        vc.border = THIN_BOTTOM
        if "IRR" in label and isinstance(value, (int, float)):
            vc.font = _irr_font(value)
            bg = _irr_bg(value)
            if bg:
                vc.fill = bg
        else:
            vc.font = F_BODY
        if alt and not getattr(vc, '_fill_applied', False):
            ws.cell(row=row, column=1).fill = LIGHT_FILL
            vc.fill = LIGHT_FILL
        row += 1

    add_ret("Equity IRR (Post-Fees, Post-MIP)", ret.irr, FMT_PCT)
    add_ret("Gross IRR (Pre-MIP)", ret.irr_gross, FMT_PCT, True)
    add_ret("Levered IRR (Pre-Fees)", ret.irr_levered, FMT_PCT)
    add_ret("Unlevered IRR", ret.irr_unlevered, FMT_PCT, True)
    row += 1
    add_ret("MOIC", ret.moic, FMT_MULT)
    add_ret("Payback (years)", ret.payback_years, "0.0", True)
    add_ret("Cash Yield (avg)", ret.cash_yield_avg, FMT_PCT)
    row += 1
    add_ret("Entry Equity", ret.entry_equity, FMT_CCY)
    add_ret("Exit Equity", ret.exit_equity, FMT_CCY, True)
    add_ret("Exit EV", ret.exit_ev, FMT_CCY)
    add_ret("Exit Net Debt", ret.exit_net_debt, FMT_CCY, True)
    add_ret("MIP Payout", ret.mip_payout, FMT_CCY)
    row += 2

    # Value bridge
    vd = state.value_drivers
    row = _write_section_header(ws, row, "VALUE CREATION BRIDGE", 3)
    ws.cell(row=row, column=1, value="Driver").font = F_HEADER
    ws.cell(row=row, column=2, value=f"{ccy}m").font = F_HEADER
    ws.cell(row=row, column=3, value="% Contribution").font = F_HEADER
    _style_header_row(ws, row, 1, 3)
    row += 1

    bridges = [
        ("Entry Equity", vd.entry_equity, None, False),
        ("(+) Revenue Growth", vd.revenue_growth_contribution_abs, vd.revenue_growth_contribution_pct, True),
        ("(+) Margin Expansion", vd.margin_expansion_contribution_abs, vd.margin_expansion_contribution_pct, False),
        ("(+) Multiple Expansion", vd.multiple_expansion_contribution_abs, vd.multiple_expansion_contribution_pct, True),
        ("(+) Debt Paydown", vd.debt_paydown_contribution_abs, vd.debt_paydown_contribution_pct, False),
        ("(-) Fees & Costs", vd.fees_drag_contribution_abs, vd.fees_drag_contribution_pct, True),
        ("Exit Equity", vd.exit_equity, None, False),
    ]
    for label, abs_val, pct_val, alt in bridges:
        is_total = label in ("Entry Equity", "Exit Equity")
        ws.cell(row=row, column=1, value=label).font = F_BODY_BOLD if is_total else F_BODY
        ws.cell(row=row, column=1).border = MED_BOTTOM if is_total else THIN_BOTTOM
        ws.cell(row=row, column=2, value=abs_val).number_format = FMT_CCY
        ws.cell(row=row, column=2).font = F_BODY_BOLD if is_total else F_BODY
        ws.cell(row=row, column=2).border = MED_BOTTOM if is_total else THIN_BOTTOM
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        if pct_val is not None:
            ws.cell(row=row, column=3, value=pct_val / 100).number_format = FMT_PCT
        ws.cell(row=row, column=3).border = MED_BOTTOM if is_total else THIN_BOTTOM
        if alt and not is_total:
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = LIGHT_FILL
        row += 1

    # EBITDA Bridge
    row += 1
    eb = state.ebitda_bridge
    if eb.entry_ebitda > 0:
        row = _write_section_header(ws, row, "EBITDA BRIDGE (Entry to Exit)", 2)
        ws.cell(row=row, column=1, value="Component").font = F_HEADER
        ws.cell(row=row, column=2, value=f"{ccy}m").font = F_HEADER
        _style_header_row(ws, row, 1, 2)
        row += 1
        eb_rows = [
            ("Entry EBITDA", eb.entry_ebitda, False),
            ("(+) Organic Revenue", eb.organic_revenue_contribution, True),
            ("(+) Margin Expansion", eb.margin_expansion_contribution, False),
            ("(+) Cost Synergies", eb.cost_synergies, True),
            ("(+) Add-On EBITDA", eb.add_on_ebitda, False),
            ("(-) Integration Costs", -eb.integration_costs, True),
            ("(-) Monitoring Fees", -eb.monitoring_fees, False),
            ("Exit EBITDA", eb.exit_ebitda, False),
        ]
        for label, val, alt in eb_rows:
            is_total = label.startswith("Entry") or label.startswith("Exit")
            ws.cell(row=row, column=1, value=label).font = F_BODY_BOLD if is_total else F_BODY
            ws.cell(row=row, column=1).border = MED_BOTTOM if is_total else THIN_BOTTOM
            ws.cell(row=row, column=2, value=val).number_format = FMT_CCY
            ws.cell(row=row, column=2).font = F_BODY_BOLD if is_total else F_BODY
            ws.cell(row=row, column=2).border = MED_BOTTOM if is_total else THIN_BOTTOM
            if alt and not is_total:
                ws.cell(row=row, column=1).fill = LIGHT_FILL
                ws.cell(row=row, column=2).fill = LIGHT_FILL
            row += 1

    # ── IRR cross-check audit ──
    row += 1
    row = _write_section_header(ws, row, "AUDIT CHECKS", 3)
    # Check 1: MOIC consistency with entry/exit equity
    moic_check = ret.exit_equity / ret.entry_equity if ret.entry_equity > 0 else 0.0
    moic_ok = abs(moic_check - ret.moic) < 0.001
    ws.cell(row=row, column=1, value="MOIC = Exit Equity / Entry Equity").font = F_GREEN if moic_ok else F_RED
    ws.cell(row=row, column=1).fill = GREEN_BG_FILL if moic_ok else RED_BG_FILL
    ws.cell(row=row, column=1).border = THIN_BOTTOM
    ws.cell(row=row, column=2, value=moic_check).number_format = FMT_MULT
    ws.cell(row=row, column=2).font = F_GREEN if moic_ok else F_RED
    ws.cell(row=row, column=2).fill = GREEN_BG_FILL if moic_ok else RED_BG_FILL
    ws.cell(row=row, column=2).border = THIN_BOTTOM
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=3, value="OK" if moic_ok else "ERROR").font = F_GREEN if moic_ok else F_RED
    ws.cell(row=row, column=3).fill = GREEN_BG_FILL if moic_ok else RED_BG_FILL
    ws.cell(row=row, column=3).border = THIN_BOTTOM
    row += 1
    # Check 2: Value bridge closes (reconciliation_delta near 0)
    bridge_ok = vd.reconciliation_delta < 0.1
    ws.cell(row=row, column=1, value="Value Bridge Reconciliation").font = F_GREEN if bridge_ok else F_RED
    ws.cell(row=row, column=1).fill = GREEN_BG_FILL if bridge_ok else RED_BG_FILL
    ws.cell(row=row, column=1).border = THIN_BOTTOM
    ws.cell(row=row, column=2, value=vd.reconciliation_delta).number_format = FMT_CCY
    ws.cell(row=row, column=2).font = F_GREEN if bridge_ok else F_RED
    ws.cell(row=row, column=2).fill = GREEN_BG_FILL if bridge_ok else RED_BG_FILL
    ws.cell(row=row, column=2).border = THIN_BOTTOM
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=3, value="OK" if bridge_ok else "GAP").font = F_GREEN if bridge_ok else F_RED
    ws.cell(row=row, column=3).fill = GREEN_BG_FILL if bridge_ok else RED_BG_FILL
    ws.cell(row=row, column=3).border = THIN_BOTTOM
    row += 1

    # ── Panel 1: Ranked Value Creation Bridge ──────────────────────────
    if vd.ranked_drivers:
        row += 1
        row = _write_section_header(ws, row, "DRIVER DECOMPOSITION (RANKED)", 3)
        ws.cell(row=row, column=1, value="Driver").font = F_HEADER
        ws.cell(row=row, column=2, value=f"{ccy}m Contribution").font = F_HEADER
        ws.cell(row=row, column=3, value="% of Equity Gain").font = F_HEADER
        _style_header_row(ws, row, 1, 3)
        row += 1
        for i, dr in enumerate(vd.ranked_drivers):
            is_neg = dr.abs_contribution < 0
            f = F_RED if is_neg else F_BODY
            alt = i % 2 == 1
            ws.cell(row=row, column=1, value=f"#{dr.rank}  {dr.name}").font = f
            ws.cell(row=row, column=1).border = THIN_BOTTOM
            if alt:
                ws.cell(row=row, column=1).fill = LIGHT_FILL
            vc = ws.cell(row=row, column=2, value=dr.abs_contribution)
            vc.number_format = FMT_CCY
            vc.font = F_RED if is_neg else F_BODY
            vc.border = THIN_BOTTOM
            vc.alignment = Alignment(horizontal="right")
            if alt:
                vc.fill = LIGHT_FILL
            pc = ws.cell(row=row, column=3, value=dr.pct_of_gain / 100)
            pc.number_format = FMT_PCT
            pc.font = F_RED if is_neg else F_BODY
            pc.border = THIN_BOTTOM
            pc.alignment = Alignment(horizontal="right")
            if alt:
                pc.fill = LIGHT_FILL
            row += 1
        # Operational vs financial split
        row += 1
        split_rows = [
            ("Operational (Revenue + Margin)", vd.operational_value_pct / 100),
            ("Financial Engineering (Multiple + Deleverage)", vd.financial_engineering_pct / 100),
        ]
        for label, pct_val in split_rows:
            ws.cell(row=row, column=1, value=label).font = F_SM_BOLD
            ws.cell(row=row, column=1).border = THIN_BOTTOM
            pv = ws.cell(row=row, column=3, value=pct_val)
            pv.number_format = FMT_PCT
            pv.font = F_SM_BOLD
            pv.border = THIN_BOTTOM
            pv.alignment = Alignment(horizontal="right")
            row += 1
        # IC insight rows
        if vd.insight_primary_driver:
            row += 1
            row = _write_section_header(ws, row, "IC INSIGHTS — DRIVER DECOMPOSITION", 3)
            for insight in [vd.insight_primary_driver, vd.insight_weak_thesis, vd.insight_overreliance_multiple]:
                if not insight:
                    continue
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                ic = ws.cell(row=row, column=1, value=insight)
                ic.font = Font(name="Calibri", size=9, color="111111", italic=True)
                ic.alignment = Alignment(wrap_text=True, vertical="top")
                ic.border = THIN_BOTTOM
                ws.row_dimensions[row].height = 32
                row += 1

    # ── Panel 2: Fragility Dashboard ───────────────────────────────────
    fa = state.fragility_analysis
    if fa.stress_scenarios:
        row += 1
        row = _write_section_header(ws, row, "FRAGILITY DASHBOARD", 3)

        # Classification badge
        cls_fill = {"Robust": GREEN_BG_FILL, "Moderate Risk": AMBER_BG_FILL, "Fragile": RED_BG_FILL}
        cls_font = {"Robust": F_GREEN, "Moderate Risk": F_AMBER, "Fragile": F_RED}
        f_fill = cls_fill.get(fa.classification, LIGHT_FILL)
        f_font = cls_font.get(fa.classification, F_BODY)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cc = ws.cell(
            row=row, column=1,
            value=f"Classification: {fa.classification}  |  Fragility Score: {fa.fragility_score:.0%}  |  Dominant Risk: {fa.dominant_stress_driver}"
        )
        cc.font = f_font
        cc.fill = f_fill
        cc.border = MED_BOTTOM
        cc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        # Stress scenario table
        ws.cell(row=row, column=1, value="Stress Scenario").font = F_HEADER
        ws.cell(row=row, column=2, value="IRR").font = F_HEADER
        ws.cell(row=row, column=3, value="ΔIRR vs Base").font = F_HEADER
        _style_header_row(ws, row, 1, 3)
        row += 1

        base_irr_str = f"{fa.base_irr:.1%}" if fa.base_irr is not None else "N/A"
        ws.cell(row=row, column=1, value=f"Base Case").font = F_BODY_BOLD
        ws.cell(row=row, column=1).border = THIN_BOTTOM
        ws.cell(row=row, column=2, value=fa.base_irr).number_format = FMT_PCT
        ws.cell(row=row, column=2).font = F_BODY_BOLD
        ws.cell(row=row, column=2).border = THIN_BOTTOM
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=3, value=0.0).number_format = FMT_PCT
        ws.cell(row=row, column=3).font = F_BODY_BOLD
        ws.cell(row=row, column=3).border = THIN_BOTTOM
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")
        row += 1

        for i, sc in enumerate(fa.stress_scenarios):
            is_combined = sc.name == "combined"
            alt = i % 2 == 0
            label = sc.description
            irr_font = F_RED if (sc.irr is not None and sc.irr < 0.12) else (F_AMBER if (sc.irr is not None and sc.irr < 0.18) else F_BODY)
            ws.cell(row=row, column=1, value=label).font = F_BODY_BOLD if is_combined else F_BODY
            ws.cell(row=row, column=1).border = MED_BOTTOM if is_combined else THIN_BOTTOM
            if alt:
                ws.cell(row=row, column=1).fill = LIGHT_FILL
            vc = ws.cell(row=row, column=2, value=sc.irr)
            vc.number_format = FMT_PCT
            vc.font = irr_font
            vc.border = MED_BOTTOM if is_combined else THIN_BOTTOM
            vc.alignment = Alignment(horizontal="right")
            if alt:
                vc.fill = LIGHT_FILL
            dc = ws.cell(row=row, column=3, value=sc.delta_irr)
            dc.number_format = FMT_PCT
            dc.font = F_RED if (sc.delta_irr is not None and sc.delta_irr < 0) else F_BODY
            dc.border = MED_BOTTOM if is_combined else THIN_BOTTOM
            dc.alignment = Alignment(horizontal="right")
            if alt:
                dc.fill = LIGHT_FILL
            row += 1

        # Fragility insight callouts
        if any([fa.insight_irr_drop, fa.insight_dominant_driver, fa.insight_classification]):
            row += 1
            row = _write_section_header(ws, row, "IC INSIGHTS — FRAGILITY", 3)
            for insight in [fa.insight_irr_drop, fa.insight_dominant_driver, fa.insight_classification]:
                if not insight:
                    continue
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                ic = ws.cell(row=row, column=1, value=insight)
                ic.font = Font(name="Calibri", size=9, color="111111", italic=True)
                ic.alignment = Alignment(wrap_text=True, vertical="top")
                ic.border = THIN_BOTTOM
                ws.row_dimensions[row].height = 32
                row += 1

    _set_print(ws, landscape=False)


# ── Sheet 7: Scenarios & Sensitivity ─────────────────────────────────────

def _build_scenarios_sheet(wb: Workbook, state: ModelState, ccy: str):
    ws = wb.create_sheet("Scenarios")
    ws.sheet_properties.tabColor = "ca6f1e"
    ws.column_dimensions["A"].width = 24

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1, value=f"{state.deal_name} - Scenario & Sensitivity Analysis").font = F_SECTION
    ws.cell(row=row, column=1).border = THICK_BOTTOM
    row += 2

    if state.scenarios:
        n = len(state.scenarios)
        row = _write_section_header(ws, row, "SCENARIO COMPARISON", n + 1)
        headers = ["Metric"] + [s.name.upper() for s in state.scenarios]
        for i, h in enumerate(headers):
            ws.cell(row=row, column=i + 1, value=h)
            ws.column_dimensions[get_column_letter(i + 1)].width = 16 if i > 0 else 24
        _style_header_row(ws, row, 1, n + 1)
        row += 1

        ws.cell(row=row, column=1, value="Equity IRR").font = F_BODY
        for i, s in enumerate(state.scenarios):
            c = ws.cell(row=row, column=i + 2, value=s.irr)
            c.number_format = FMT_PCT
            c.font = _irr_font(s.irr)
            bg = _irr_bg(s.irr)
            if bg:
                c.fill = bg
            c.alignment = Alignment(horizontal="center")
        row += 1

        ws.cell(row=row, column=1, value="MOIC").font = F_BODY
        for i, s in enumerate(state.scenarios):
            c = ws.cell(row=row, column=i + 2, value=s.moic)
            c.number_format = FMT_MULT
            c.alignment = Alignment(horizontal="center")
        row += 1

        ws.cell(row=row, column=1, value="Exit Multiple").font = F_BODY
        for i, s in enumerate(state.scenarios):
            c = ws.cell(row=row, column=i + 2, value=s.exit_multiple)
            c.number_format = FMT_MULT
            c.alignment = Alignment(horizontal="center")
        row += 1

        ws.cell(row=row, column=1, value=f"Exit Equity ({ccy}m)").font = F_BODY
        for i, s in enumerate(state.scenarios):
            c = ws.cell(row=row, column=i + 2, value=s.exit_equity)
            c.number_format = FMT_CCY
            c.alignment = Alignment(horizontal="center")
        row += 3

    # Sensitivity tables
    if state.sensitivity_tables:
        row = _write_section_header(ws, row, "SENSITIVITY ANALYSIS - IRR", 10)
        row += 1

        labels = {
            "revenue_growth": "Revenue Growth",
            "exit_multiple": "Exit Multiple",
            "ebitda_margin": "EBITDA Margin",
            "entry_multiple": "Entry Multiple",
            "leverage": "Leverage",
        }

        for table in state.sensitivity_tables:
            rl = labels.get(table.row_variable, table.row_variable)
            cl = labels.get(table.col_variable, table.col_variable)
            nc = len(table.col_values)

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc + 1)
            ws.cell(row=row, column=1, value=f"Table {table.table_id}: {rl} vs {cl}").font = F_BODY_BOLD
            ws.cell(row=row, column=1).border = MED_BOTTOM
            row += 1

            is_col_pct = table.col_variable in ("revenue_growth", "ebitda_margin")
            is_row_pct = table.row_variable in ("revenue_growth", "ebitda_margin")

            ws.cell(row=row, column=1, value=f"{rl} \\ {cl}")
            for j, cv in enumerate(table.col_values):
                c = ws.cell(row=row, column=j + 2, value=cv)
                c.number_format = FMT_PCT if is_col_pct else FMT_MULT
                ws.column_dimensions[get_column_letter(j + 2)].width = 12
            _style_header_row(ws, row, 1, nc + 1)
            row += 1

            for i, rv in enumerate(table.row_values):
                rhc = ws.cell(row=row, column=1, value=rv)
                rhc.number_format = FMT_PCT if is_row_pct else FMT_MULT
                rhc.font = F_BODY_BOLD
                rhc.fill = MID_FILL
                rhc.border = THIN_BOTTOM

                for j in range(nc):
                    irr = table.irr_matrix[i][j] if i < len(table.irr_matrix) and j < len(table.irr_matrix[i]) else None
                    c = ws.cell(row=row, column=j + 2, value=irr)
                    c.number_format = FMT_PCT
                    c.font = _irr_font(irr)
                    c.alignment = Alignment(horizontal="center")
                    c.border = THIN_BOTTOM
                    bg = _irr_bg(irr)
                    if bg:
                        c.fill = bg
                    elif i % 2 == 1:
                        c.fill = LIGHT_FILL
                row += 1
            row += 2

    _set_print(ws)


# ── Sheet 8: Risk Assessment ────────────────────────────────────────────

def _build_risk_sheet(wb: Workbook, state: ModelState, ccy: str, hp: int):
    ws = wb.create_sheet("Risk")
    ws.sheet_properties.tabColor = "c0392b"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 40

    rc = state.exit_reality_check
    ca = state.credit_analysis
    row = 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=f"{state.deal_name} - Risk Assessment").font = F_SECTION
    ws.cell(row=row, column=1).border = THICK_BOTTOM
    row += 2

    # Verdict banner
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    vc = ws.cell(row=row, column=1, value=f"EXIT REALITY CHECK: {rc.verdict.upper()}")
    vc.font = F_WHITE_LG
    vc.fill = _verdict_fill(rc.verdict)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 2

    row = _write_section_header(ws, row, "EXIT VALUATION METRICS", 2)
    row = _write_kv_row(ws, row, "Exit EV / EBITDA", rc.ev_ebitda_at_exit, fmt=FMT_MULT)
    row = _write_kv_row(ws, row, "Exit EV / Revenue", rc.ev_revenue_at_exit, fmt=FMT_MULT, alt=True)
    row = _write_kv_row(ws, row, "Multiple Delta", rc.multiple_delta, fmt='0.0"x"')
    lo, hi = rc.public_comps_multiple_range
    row = _write_kv_row(ws, row, "Sector Comps Range", f"{lo:.1f}x - {hi:.1f}x", alt=True)
    if rc.implied_buyer_irr is not None:
        row = _write_kv_row(ws, row, "Implied Buyer IRR", rc.implied_buyer_irr, fmt=FMT_PCT, font=_irr_font(rc.implied_buyer_irr))
    row += 1

    if rc.flags:
        row = _write_section_header(ws, row, "RISK FLAGS", 4)
        for col, val in [(1, "Rule"), (2, "Severity"), (3, "Description"), (4, "Quantified Impact")]:
            ws.cell(row=row, column=col, value=val)
        _style_header_row(ws, row, 1, 4)
        row += 1

        for i, flag in enumerate(rc.flags):
            alt = i % 2 == 1
            name = flag.flag_type.replace("_", " ").title()
            ws.cell(row=row, column=1, value=name).font = F_BODY
            ws.cell(row=row, column=1).border = THIN_BOTTOM

            sev = ws.cell(row=row, column=2, value=flag.severity.upper())
            sev.font = F_RED if flag.severity == "critical" else F_AMBER
            sev.fill = RED_BG_FILL if flag.severity == "critical" else AMBER_BG_FILL
            sev.alignment = Alignment(horizontal="center")
            sev.border = THIN_BOTTOM

            ws.cell(row=row, column=3, value=flag.description).font = F_BODY
            ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=3).border = THIN_BOTTOM

            ws.cell(row=row, column=4, value=flag.quantified_impact).font = F_BODY_BOLD
            ws.cell(row=row, column=4).border = THIN_BOTTOM

            if alt:
                ws.cell(row=row, column=1).fill = LIGHT_FILL
                ws.cell(row=row, column=3).fill = LIGHT_FILL
                ws.cell(row=row, column=4).fill = LIGHT_FILL
            ws.row_dimensions[row].height = 30
            row += 1
        row += 1

    # Credit analysis
    if ca.metrics_by_year:
        row = _write_section_header(ws, row, "CREDIT ANALYSIS", 2)
        row = _write_kv_row(ws, row, "Estimated Credit Rating", ca.credit_rating_estimate)
        row = _write_kv_row(ws, row, f"Max Capacity @ 4x ({ccy}m)", ca.max_debt_capacity_at_4x, fmt=FMT_CCY, alt=True)
        row = _write_kv_row(ws, row, f"Max Capacity @ 5x ({ccy}m)", ca.max_debt_capacity_at_5x, fmt=FMT_CCY)
        row = _write_kv_row(ws, row, f"Max Capacity @ 6x ({ccy}m)", ca.max_debt_capacity_at_6x, fmt=FMT_CCY, alt=True)
        row += 1

        ws.cell(row=row, column=1, value="Refinancing Risk").font = F_BODY_BOLD
        rfc = ws.cell(row=row, column=2, value="YES" if ca.refinancing_risk else "NO")
        rfc.font = F_RED if ca.refinancing_risk else F_GREEN
        rfc.fill = RED_BG_FILL if ca.refinancing_risk else GREEN_BG_FILL
        rfc.alignment = Alignment(horizontal="center")
        row += 1
        if ca.refinancing_risk_detail:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1, value=ca.refinancing_risk_detail).font = F_SM_BOLD
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1
        row += 1

        if ca.recovery_waterfall:
            row = _write_section_header(ws, row, "RECOVERY WATERFALL (50% EV Stress)", 2)
            ws.cell(row=row, column=1, value="Tranche")
            ws.cell(row=row, column=2, value="Recovery %")
            _style_header_row(ws, row, 1, 2)
            row += 1
            for i, rw in enumerate(ca.recovery_waterfall):
                ws.cell(row=row, column=1, value=rw.tranche).font = F_BODY
                ws.cell(row=row, column=1).border = THIN_BOTTOM
                rc_cell = ws.cell(row=row, column=2, value=rw.recovery_pct)
                rc_cell.number_format = FMT_PCT
                rc_cell.font = F_GREEN if rw.recovery_pct >= 1 else (F_AMBER if rw.recovery_pct >= 0.5 else F_RED)
                rc_cell.border = THIN_BOTTOM
                rc_cell.alignment = Alignment(horizontal="center")
                if i % 2 == 1:
                    ws.cell(row=row, column=1).fill = LIGHT_FILL
                    rc_cell.fill = LIGHT_FILL
                row += 1

    _set_print(ws, landscape=False)
